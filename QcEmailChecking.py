#latest working file on 9th april 2021
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from string import Template
import mimetypes
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
import json
import imaplib
from datetime import date ,datetime, timedelta
import datetime
import email
from email.message import EmailMessage
import os
import sys
import os.path
from os import path
from os import listdir
from os.path import isfile,join
import base64
import smtplib
import mimetypes
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
import logging
from imap_tools import MailBox, A
import schedule
import time

root = Path(__file__).parent.resolve()

logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s %(levelname)s %(message)s',
                    filename=os.path.join(root,"app.log"),
                    filemode='a')

class EmailChecking():
    def __init__(self):
        print("Starting Qc mail checking bot")
        with open(os.path.join(root, 'botDetails.json'), 'r') as file:
       

            data = json.load(file)

            for x in data['bot']:
                self.errorReportemail = x['erroremail']
                self.errorReportemailpasswd = x['errorpwd']
                self.aocEmail = x['AOCEmail'] 
                self.aocPwd = x ['AOCpwd']
                self.aocError1 = x ['AOCError1']
                self.aocError2 = x ['AOCError2']
               
        self.get_details()
        logging.info('Qc email bot starting')

    def get_details(self):
        try:
            with open(os.path.join(root, 'loginHotel.json')) as json_file:

                data = json.load(json_file)
                for p in data['hotels']:
                    groupname = p['hotelGroupName']
                    
                    loginEmail = p['loginEmails']
                    password = p['loginPasswords']

                    try:

                        imapSession =  imaplib.IMAP4_SSL("imap.gmail.com")
                        imapSession.login(loginEmail,password)
                    except Exception as e:
                        print('81')
                        print (e)
                        logging.info('Error Login to email ' + loginEmail)
                        subject = 'Qc Email BOT - Error Login to:{}'.format(loginEmail)
                        message =(f"""Qc Email BOT - Error Login to: {loginEmail}
            App: Qc Email BOT   
            Reason: Error Login to above email 
            Fix: Check JSON file
            [THIS IS AN AUTOMATED MESSAGE - PLEASE DO NOT REPLY DIRECTLY TO THIS EMAIL]
                        """)
                    
                        
                        try:
                           self.sendErrorEmail(self.aocEmail,self.aocPwd,self.aocError1,self.aocError2,subject,message)
                        except Exception as e:
                                print('Error while Sending failure email \nError Details: {}'.format(e))
                                logging.warning('Error while Sending QC email \nError Details: {}'.format(e))

                        continue
                   
                    
                    logging.info("Checking hotel details for " + loginEmail)
                    try:
                       
                        for r in p['hotel']:

                            hotelname = r['hotelName']
                            reports = r['reports']
                            print ("HMG Name" + groupname)
                            print("hotel name " + hotelname)
                           
                        
                            for x in r['reports']:
                                subject = x['subject']
                                mailfrom = x['mailfrom']
                                

                                self.checkAttachement(loginEmail,password,subject,mailfrom,groupname,hotelname)
                                
           
                        for m in data['qcmail']:
                            qcemail = m['email']
                            # ccmail = m['ccmail']
                           

                            self.sendMailAtachment(groupname,hotelname,self.errorReportemail,self.errorReportemailpasswd,qcemail)
                        #  logger.exception("FATAL ERROR: {}".format(e), exc_info=False)   
                    except Exception as e:
                        print('127')
                        print(e)
                        logging.info("Error Reading hotel details of  "+ loginEmail)
                       
                        logging.info("Error Reading hotel details of  :{}".format(e))
                        subject = 'Qc Email BOT - Error Reading hotel details of:{}'.format(e)
                        message =(f"""Qc Email BOT - Error Reading hotel details of: {loginEmail}
App: Qc Email BOT   
Reason: Error Reading hotel details 
Fix: Check JSON file
[THIS IS AN AUTOMATED MESSAGE - PLEASE DO NOT REPLY DIRECTLY TO THIS EMAIL]
            """)        
                        try:
                            self.sendErrorEmail(self.aocEmail,self.aocPwd,self.aocError1,self.aocError2,subject,message)

                        except Exception as e:
                            print(e)
                            print('Error while Sending failure email \nError Details: {}'.format(e))
                            logging.warning('Error while Sending QC email \nError Details: {}'.format(e))

                        continue
                        
        except Exception as e:
            print(e)

  
       
    def checkDirectory(self):

        try:
            businessdate = date.today() - timedelta(days=1)
            directory = str(businessdate)
            path = os.path.join(root,directory)
            isFile = os.path.isdir(path)

            if isFile == True:
                return path
            else:
                os.mkdir(path)
                return path
                
                # print(path)
        except Exception as e:
            print(e)
                

    def generatingFileName(self,groupname):

        try:
            filedir = self.checkDirectory()
            
            businessdate = date.today() - timedelta(days=1)
            filename = groupname + str (businessdate) +".xlsx"
            path = os.path.join(filedir,filename)
            return path
        except Exception as e:
            print(e)
            logging.info("Error In Generating Folder  :{}".format(e))
    def writingToExcel(self,loginEmail,date,subject,mailfrom,reportExist,reportType,groupname,hotelname):
            try:
                businessdate = (datetime.date.today() - datetime.timedelta(1)).strftime("%d-%m-%Y")
                
                report = reportType
                path = self.generatingFileName(groupname)
                
                if os.path.isfile(path):

                    
                    global  wb_obj 
                    wb_obj= openpyxl.load_workbook(path)
                   
                    
                    if hotelname in wb_obj.sheetnames:
                       
                        sheet = wb_obj[hotelname]
                        sheet.cell(column=1, row=sheet.max_row+1, value=loginEmail)
                        sheet.cell(column=2, row=sheet.max_row, value=businessdate)
                        sheet.cell(column=3, row=sheet.max_row, value=subject)
                        sheet.cell(column=4, row=sheet.max_row, value=mailfrom)
                        sheet.cell(column=5, row=sheet.max_row, value=reportExist)
                        sheet.cell(column=6, row=sheet.max_row, value=report)

                        wb_obj.save(path)

                    else:
                       
                        sheet = wb_obj.create_sheet()
                       
                        sheet.title = hotelname
                        sheet["A1"] = "Receiver"
                        sheet["B1"] = "Business Date"
                        sheet["C1"] = "PMS Report Name"
                        sheet["D1"] = "Email"
                        sheet["E1"] = "Exist Or Not Exist"
                        sheet["F1"] = "Report Type"

                        sheet["A2"] =loginEmail
                        sheet["B2"] = businessdate
                        sheet["C2"] = subject
                        sheet["D2"] = mailfrom
                        sheet["E2"] = reportExist
                        sheet["F2"] = report


                        wb_obj.save(path)
                   
                    
                else:
                    global wb
                    wb = openpyxl.Workbook()
                    
                    sheet = wb.create_sheet()
                    

                    sheet.title = hotelname
                    sheet["A1"] = "Receiver"
                    sheet["B1"] = "Business Date"
                    sheet["C1"] = "PMS Report Name"
                    sheet["D1"] = "Email"
                    sheet["E1"] = "Exist Or Not Exist"
                    sheet["F1"] = "Report Type"

                    sheet["A2"] =loginEmail
                    sheet["B2"] = businessdate
                    sheet["C2"] = subject
                    sheet["D2"] = mailfrom
                    sheet["E2"] = reportExist
                    sheet["F2"] = report

                    wb.save(path)

            except Exception as e:
                logging.info("Error In Generating Excel  :{}".format(e))
                

    # self.checkAttachement(loginEmail,password,subject,mailfrom,groupname,hotelname)

    def checkAttachement(self,loginEmail,password,subject,mailfrom,groupname,hotelname):
        imapSession =  imaplib.IMAP4_SSL("imap.gmail.com")
        try:

            imapSession.login(loginEmail,password)
            try:

                typ, data = imapSession.select('"[Gmail]/All Mail"')
                    # date = (datetime.date.today() - datetime.timedelta(1)).strftime("%d-%b-%Y")
                date = (datetime.date.today() - datetime.timedelta(1)).strftime("%d-%b-%Y")
                typ, data = imapSession.search(None, '(SENTON "'+ date + '")','(SUBJECT "'+ subject + '")','(TO "'+ mailfrom + '")')
                
                checkForAttachement = data[0].split()
            
                if ((len(checkForAttachement)) == 0):
                    print('286')
                    reportExist = 'No email'
                    reportType = 'N/A'
                    print('no email')
                    self.writingToExcel(loginEmail,date,subject,mailfrom,reportExist,reportType,groupname,hotelname)
                    
                else:
                    print('314')
                    with MailBox('imap.gmail.com').login(loginEmail,password,initial_folder='[Gmail]/All Mail') as mailbox:
                        date = (datetime.date.today() - datetime.timedelta(1)).strftime("%d-%b-%Y")
                        print('318')
                        # on
                        for msg in mailbox.fetch(A('(ON "'+ date + '")','(SUBJECT "'+ subject + '")','(TO "'+ mailfrom + '")')):
                            print('321')
                            print(msg.date)
                            print(msg.subject)
                            print(msg.attachments)
                            if any(att.filename.lower().endswith('.csv') for att in msg.attachments):
                                print('- has csv')
                                reportExist = 'CSV Exsist'
                                reportType = 'CSV'
                                self.writingToExcel(loginEmail,date,msg.subject,mailfrom,reportExist,reportType,groupname,hotelname)

                            elif any(att.filename.lower().endswith('.xlsx') for att in msg.attachments):
                                print('- has xlsx')
                                reportExist = 'xlsx Not Exsist'
                                reportType = 'XLSX'
                                self.writingToExcel(loginEmail,date,msg.subject,mailfrom,reportExist,reportType,groupname,hotelname)
                            # elif any(att.filename.lower().endswith('.txt') for att in msg.attachments):
                            #     print('- has txt')
                            #     reportExist = 'PDF Not Exsist'
                            #     reportType = 'TXT'
                            #     self.writingToExcel(loginEmail,date,msg.subject,mailfrom,reportExist,reportType,groupname,hotelname)
                            else:
                                reportExist = 'PDF Not Exsist'
                                reportType = 'N/A'
                                print('N/A')
                                self.writingToExcel(loginEmail,date,msg.subject,mailfrom,reportExist,reportType,groupname,hotelname)
            except Exception as e:
                print(e)
                
                logging.info("Error checking attachement :{}".format(e))  
                subject = 'Qc Email BOT - Error Reading hotel details of:{}'.format(e)
                message =(f"""Qc Email BOT - Error Reading hotel details of: {loginEmail}
App: Qc Email BOT   
Reason: Error Reading hotel details 
Fix: Check JSON file
[THIS IS AN AUTOMATED MESSAGE - PLEASE DO NOT REPLY DIRECTLY TO THIS EMAIL]
            """)
                try:
                    self.sendErrorEmail(self.aocEmail,self.aocPwd,self.aocError1,self.aocError2,subject,message)
                except Exception as e:
                    print('Error while Sending failure email \nError Details: {}'.format(e))
                    logging.warning('Error while Sending QC email \nError Details: {}'.format(e))


                

        except Exception as e:
            print(e)
            
    
#  self.sendMailAtachment(groupname,hotelname,self.errorReportemail,self.errorReportemailpasswd,qcemail)
    def sendMailAtachment(self,groupname,hotelname,email,password,qcemail):
        try:
            imapSession =  smtplib.SMTP(host = "smtp.gmail.com",port = 587)
            imapSession.starttls()
        
            subject = "Email QC Bot for   " +groupname     
                
            imapSession.login(email,password)
            msg = MIMEMultipart()
            msg['Subject'] = subject
            
            msg['From'] = email
            msg['To'] = qcemail
                
            path = self.checkDirectory()
            
            onlyfiles =[f for f in listdir(path) if isfile(join(path,f))]
            for filexl in onlyfiles:
            
                filexlwithpath = join(path,filexl)
                
                fp = open(filexlwithpath,'rb')

                filename = filexl
                print(filename)
            
                xls = MIMEBase('application','vnd.ms-excel')
                xls.set_payload(fp.read())
                fp.close()
                encoders.encode_base64(xls)
                xls.add_header('Content-Disposition', 'attachment', filename=filename)
                msg.attach(xls)
                
                filexlwithpath = join(path,filexl)
                try:

            
                    imapSession.send_message(msg)
                except Exception as e:
                    print(e)
                else:
                    os.remove(join(path,filexl))
                    print('removed '+ filexl)
                    logging.info('Attachement Removed ' + filexl)

                # imapSession.close()
                # print('sent')
                logging.info('Attachement Sent ' + filexl)
        except Exception as e:
            print(e)
            logging.info("Error Sending attachment {}".format(e))
            # 'Qc Email BOT - Error Sending Attachment hotel details of:{}'.format(e)
            subject = 'Qc Email BOT - Error Sending Attachment hotel details of:{}'.format(e, groupname)
            message =(f"""Qc Email BOT - Error Sending Attachment
App: Qc Email BOT   
Reason: sending attachment 
Fix: Check login details of QC Automation BOT
[THIS IS AN AUTOMATED MESSAGE - PLEASE DO NOT REPLY DIRECTLY TO THIS EMAIL]
            """)
            try:
                self.sendErrorEmail(self.aocEmail,self.aocPwd,self.aocError1,self.aocError2,subject,message)
            except Exception as e:
                print('Error while Sending failure email \nError Details: {}'.format(e))
                logging.warning('Error while Sending failure email \nError Details: {}'.format(e))


    def sendErrorEmail(self, email, password, send_to_1,send_to_2 ,subject, message):
        print("Sending email to", send_to_1)
        logging.info("Sending error email")
        try:
            
            msg = MIMEMultipart()
            # recipients = [send_to_1,send_to_2]
            msg["From"] = email
            # msg["To"] = ", ".join(recipients)
            msg["To"] = send_to_1
            msg["Cc"] = send_to_2
        
            msg["Subject"] = subject
            msg.attach(MIMEText(message, 'plain'))
            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()
            server.login(email, password)
            text = msg.as_string()
            server.sendmail(email, send_to_1, text)
            server.quit()
        except Exception as e:
            print('Error while Sending failure email \nError Details: {}'.format(e))
            logging.warning('Error while Sending failure email \nError Details: {}'.format(e))


            


        


def main():
    EmailChecking()  # create instance of QcEmailChecking
    
if __name__ == '__main__':
    # self.errorReportemail
    
    # schedule.every(1).minutes.do(main)
    with open(os.path.join(root, 'loginHotel.json')) as json_file:
        print('nushra')
        data = json.load(json_file)
        scheduleTime=data['sheduledTime']
        print(scheduleTime)
    

        
        schedule.every(1).day.at(scheduleTime).do(main)
        # schedule.every(1).minutes.do(main)

    while True:
        schedule.run_pending()
        time.sleep(1)
        # main()


